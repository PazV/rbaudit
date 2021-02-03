#!/usr/bin/env python
#--*-- coding: utf-8 --*--
from db_connection import getDB
db = getDB()
import os
import logging
import app_config as cfg
import sys
import traceback
import datetime
from dateutil.relativedelta import relativedelta
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Alignment, Side, NamedStyle
from openpyxl.cell import Cell
from time import gmtime, strftime, localtime

print "Entra crear form para descarga en excel"
# create logger with 'spam_application'
logger = logging.getLogger('Download form')
logger.setLevel(logging.DEBUG)
# create file handler which logs even debug messages
fh = logging.FileHandler('%sdownload_form.log'%cfg.log_path)
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.ERROR)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
# add the handlers to the logger
if not len(logger.handlers):
    logger.addHandler(fh)
    logger.addHandler(ch)

logger.info("Log info")

class GenFunc:
    def as_text(self,value):
        if value is None:
            return ""
        else:
            try:
                return str(value)
            except:
                return value

def main():
    try:
        response={}
        GF=GenFunc()
        data={
            'project_id':82 ,
            'form_id':427
        }
        project=db.query("""
            select
                name,
                company_name,
                (select a.name from system.user a where a.user_id=manager) as manager,
                (select a.name from system.user a where a.user_id=partner) as partner,
                comments
            from
                project.project
            where project_id=%s
        """%data['project_id']).dictresult()[0]
        form=db.query("""
            select
                name,
                columns_number,
                rows,
                status_id,
                columns,
                (select a.name from system.user a where a.user_id=assigned_to) as assigned_to,
                to_char(resolved_date,'DD/MM/YYYY HH24:MI:SS') as resolved_date,
                to_char(first_revision_date,'DD/MM/YYYY HH24:MI:SS') as first_revision_date
                --,revisions
            from
                project.form
            where
                form_id=%s
        """%data['form_id']).dictresult()[0]
        wb = Workbook()
        ws = wb.create_sheet('Hoja1',0)

        company_style=Font(
            name='Times New Roman',
            size=14,
            bold=True,
            italic=False,
            color='FF000080'
        )
        project_style=Font(
            name='Times New Roman',
            size=14,
            bold=False,
            italic=True,
            color='FF000080'
        )
        formname_style=Font(
            name='Times New Roman',
            size=14,
            bold=False,
            italic=False,
            color='FF000080'
        )

        header_style=NamedStyle(name='header_style')
        header_style.font=Font(
            name='Times New Roman',
            size=14,
            bold=True,
            italic=False,
            color='FFFFFFFF'
        )
        header_style.fill=PatternFill("solid", fgColor="FF000099")
        header_style.alignment=Alignment(horizontal='center')
        header_style.border=Border(
            left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
        )

        content_style=NamedStyle(name='content_style')
        content_style.font=Font(
            name='Times New Roman',
            size=12,
            bold=False,
            italic=False,
            color='FF000000'
        )
        content_style.alignment=Alignment(
            horizontal='justify',
            vertical='center',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0)
        # wrap_text=True,shrink_to_fit=False,vertical='justify',horizontal='left')
        content_style.border=Border(
            left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
        )

        ws.sheet_view.showGridLines = False #ocultar líneas

        ws['B2']=project['company_name']
        ws['B2'].font=company_style
        ws['B3']=project['name']
        ws['B3'].font=project_style
        ws['B4']=form['name']
        ws['B4'].font=formname_style

        columns=eval(form['columns'])
        col_num=2
        for c in columns:
            ws.cell(column=col_num,row=6,value=c['name'])
            ws.cell(column=col_num,row=6).style=header_style
            col_num+=1
        ws.cell(column=col_num,row=6,value='Revisión')
        ws.cell(column=col_num,row=6).style=header_style

        form_info=db.query("""
            select * from form.project_%s_form_%s order by entry_id
        """%(data['project_id'],data['form_id'])).dictresult()
        row=7

        for x in form_info:
            keys=sorted(x.iteritems())
            col_num=2
            for k in keys:
                if k[0].split('_')[0]=='col':
                    ws.cell(column=col_num,row=row,value=k[1].decode('utf-8'))
                    ws.cell(column=col_num,row=row).style=content_style

                    col_num+=1
            ws.cell(column=col_num,row=row,value=x['rev_1'].decode('utf-8'))
            ws.cell(column=col_num,row=row).style=content_style
            row+=1

        last_info_row=row

        #comprueba si el status es cerrado para agregar datos de quien realizó y revisó el formulario
        # if int(form['status_id'])==7:
        if True:


            #agregar quién realizó el formulario
            revisions=db.query("""
                select (select a.name from system.user a where a.user_id=b.user_id) as user_name,
                b.revision_number, to_char(b.revision_date,'DD/MM/YYYY HH24:MI:SS') as revision_date
                from project.form_revisions b where b.form_id=%s order by b.revision_number asc
            """%data['form_id']).dictresult()
            bd = Side(style='thick', color="000000")
            th = Side(style='thin', color="000000")

            col_num+=2
            ws.cell(column=col_num,row=6).border=Border(left=bd,top=bd,bottom=th)
            ws.cell(column=col_num+1,row=6,value='Nombre')
            ws.cell(column=col_num+1,row=6).font=Font(name='Times New Roman', size=12, bold=True)
            ws.cell(column=col_num+1,row=6).border=Border(top=bd,bottom=th)
            ws.cell(column=col_num+2,row=6,value='Fecha')
            ws.cell(column=col_num+2,row=6).font=Font(name='Times New Roman', size=12, bold=True)
            ws.cell(column=col_num+2,row=6).border=Border(top=bd,right=bd,bottom=th)

            ws.cell(column=col_num,row=7,value='Encargado')
            ws.cell(column=col_num,row=7).font=Font(name='Times New Roman', size=12, bold=True)
            ws.cell(column=col_num,row=7).border=Border(left=bd)
            ws.cell(column=col_num+1,row=7,value=form['assigned_to'])
            ws.cell(column=col_num+1,row=7).font=Font(name='Times New Roman', size=12, bold=False)
            ws.cell(column=col_num+2,row=7,value=form['resolved_date'])
            ws.cell(column=col_num+2,row=7).font=Font(name='Times New Roman', size=12, bold=False)
            ws.cell(column=col_num+2,row=7).border=Border(right=bd)

            current_row=8
            for r in revisions:
                ws.cell(column=col_num,row=current_row,value='Revisor')
                ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
                ws.cell(column=col_num,row=current_row).border=Border(left=bd)
                # revisor=db.query("""
                #     select name from system.user where user_id=%s
                # """%int(form['revisions'].split(",")[0].split(":")[1])).dictresult()[0]
                ws.cell(column=col_num+1,row=current_row,value=r['user_name'])
                ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                ws.cell(column=col_num+2,row=current_row,value=r['revision_date'])
                ws.cell(column=col_num+2,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                ws.cell(column=col_num+2,row=current_row).border=Border(right=bd)
                current_row+=1

            ws.cell(column=col_num,row=current_row,value='Gerente')
            ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
            ws.cell(column=col_num,row=current_row).border=Border(left=bd)
            ws.cell(column=col_num+1,row=current_row,value=project['manager'])
            ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
            ws.cell(column=col_num+2,row=current_row).border=Border(right=bd)
            current_row+=1

            ws.cell(column=col_num,row=current_row,value='Socio')
            ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
            ws.cell(column=col_num,row=current_row).border=Border(left=bd,bottom=bd)
            ws.cell(column=col_num+1,row=current_row).border=Border(bottom=bd)
            ws.cell(column=col_num+1,row=current_row,value=project['partner'])
            ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
            ws.cell(column=col_num+2,row=current_row).border=Border(right=bd,bottom=bd)


            for column_cells in ws.columns:
                length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column].width = length


            #revisar si hay observaciones del formulario
            comments=db.query("""
                select b.comment, to_char(b.created,'DD/MM/YYYY HH24:MI:SS') as created, (select a.name from system.user a where a.user_id=b.user_id) as user_name from project.form_comments b where b.form_id=%s
            """%data['form_id']).dictresult()
            if comments!=[]:
                comm_header_style=NamedStyle(name='comm_header_style')
                comm_header_style.font=Font(
                    name='Times New Roman',
                    size=14,
                    bold=True,
                    italic=False,
                    color='FFFFFFFF'
                )
                comm_header_style.fill=PatternFill("solid", fgColor="FF7082D4")
                comm_header_style.alignment=Alignment(horizontal='center')
                comm_header_style.border=Border(
                    left=Side(border_style='thin',color='FF000000'),
                    right=Side(border_style='thin',color='FF000000'),
                    top=Side(border_style='thin',color='FF000000'),
                    bottom=Side(border_style='thin',color='FF000000')
                )

                wo = wb.create_sheet('Observaciones',1)
                row=2
                wo.cell(column=2, row=row, value='Por')
                wo.cell(column=2, row=row).style=comm_header_style
                wo.cell(column=3, row=row, value='Fecha')
                wo.cell(column=3, row=row).style=comm_header_style
                wo.cell(column=4, row=row, value='Observación')
                wo.cell(column=4, row=row).style=comm_header_style
                row+=1
                for x in comments:
                    wo.cell(column=2,row=row,value=x['user_name'])
                    wo.cell(column=2,row=row).style=content_style
                    wo.cell(column=3,row=row,value=x['created'])
                    wo.cell(column=3,row=row).style=content_style
                    wo.cell(column=4,row=row,value=x['comment'])
                    wo.cell(column=4,row=row).style=content_style
                    row+=1

                for column_cells in wo.columns:
                    length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                    wo.column_dimensions[column_cells[0].column].width = length





        #Descarga de archivo
        time=strftime("%H_%M_%S", gmtime())
        path=os.path.join(cfg.uploaded_forms_files_path,'Reporte%s.xlsx'%time)
        wb.save(path)
        response['success']=True
        response['msg_response']='El formulario ha sido generado.'
        response['filename']='/project/downloadFile/report/Reporte%s.xlsx'%time
        logger.info(response)
    except:
        exc_info = sys.exc_info()
        logger.error(traceback.format_exc(exc_info))

if __name__ == '__main__':
    main()
